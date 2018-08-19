import openpyxl

loc = ("Book1.xlsx")
wb = openpyxl.load_workbook(loc)
sheet = wb.get_sheet_by_name('Sheet1')
n = sheet.max_row
print n

for i in range(1, n):

    xml = """<?xml version="1.0" encoding="UTF-8"?>
<Freightcom xmlns="http://www.freightcom.net/XMLSchema" username="fcomtest12345" password="fcomtest12345" version="3.1.0">
    <QuoteRequest serviceId="0" stackable="true">
        <From id="123" company='""" + sheet.cell(row=i, column=1).value + """' address1='""" + sheet.cell(row=i, column=2).value + """' city='""" + sheet.cell(row=i, column=3).value + """' state='""" + sheet.cell(row=i, column=4).value + """' country='CA' zip='""" + sheet.cell(row=i, column=5).value + """' />
        <To company="Test Company" address1="650 CIT Drive" city="Livingston" state="ON" zip="L4J7Y9" country="CA" />
        <Packages type="Package">
        <!-- Description is ignore for quote requests -->
        <!-- If package type="Pallet", type and freightClass are required for quote requests -->
            <Package length="15" width="10" height="12" weight="10" type="Pallet" freightClass="70" nmfcCode="XXXX56" insuranceAmount="0.0" codAmount="0.0" description="desc."/>
            <Package length="15" width="10" height="10" weight="5" type="Pallet" freightClass="70" insuranceAmount="0.0" codAmount="0.0" description="desc."/>
        </Packages>
    </QuoteRequest>
</Freightcom>"""

    # pp = pprint.PrettyPrinter(indent=4)
    # pp.pprint(json.dumps(xmltodict.parse(r.content)))
    with open('quotes.xml') as fd:
        doc = xmltodict.parse(fd.read())
    fd.close

    myobj = json.loads(json.dumps(doc))
    #dictionary = ast.literal_eval(myobj['Quotes'])
    #for key, nested in dictionary.iteritems():
    #    nested = ast.literal_eval(nested)
    #    print key, nested
    #jdata = ast.literal_eval(json.dumps(doc))
    #pp = pprint.PrettyPrinter(indent=4)
    #pp.pprint(json.dumps(doc))
    # js = json.dumps(doc)
    # print js
    # js1 = (js.encode('ascii', 'ignore')).decode("utf-8")
    # print js1
    #js2 = json.loads(str(jdata))

    #print js2
    #js3 = (js2.encode('ascii', 'ignore')).decode("utf-8")
    #print js3
    #with open ('quotes.json' , 'wb') as fj:
    #    fj.write(json.dumps(doc))
    #fj.close()
    #pandas.read_json(json.dumps(doc)).to_excel("output.xlsx")

    print xml