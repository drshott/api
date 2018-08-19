import requests
from xml.etree import ElementTree as ET
import openpyxl
import xlwt

loc = ("Book1.xlsx")
wb = openpyxl.load_workbook(loc)
sheet = wb['Sheet1']
book = xlwt.Workbook("Book1.xlsx")
sh = book.get_active_sheet
n = sheet.max_row + 1

for i in range(2, n):

    xml = """<?xml version="1.0" encoding="UTF-8"?>
        <Freightcom xmlns="http://www.freightcom.net/XMLSchema" username="freightcomcost" password="1234" version="3.1.0">
            <QuoteRequest serviceId="0" stackable="true">
                <From id="123" company='""" + sheet.cell(row=i, column=1).value + """' address1='""" + sheet.cell(row=i, column=2).value + """' city='""" + sheet.cell(row=i, column=3).value + """' state='""" + sheet.cell(row=i, column=4).value + """' country='CA' zip='""" + sheet.cell(row=i, column=5).value + """' />
                <To company="Test Company" address1="650 CIT Drive" city="Livingston" state="ON" zip="L4J7Y9" country="CA" />
                <Packages type="Package">
                <!-- Description is ignore for quote requests -->
                <!-- If package type="Pallet", type and freightClass are required for quote requests -->
                    <Package length="15" width="10" height="12" weight="200" type="Package" nmfcCode="XXXX56" insuranceAmount="0.0" codAmount="0.0" description="desc."/>
                </Packages>
            </QuoteRequest>
        </Freightcom>"""

    headers = {'Content-Type': 'application/xml'}
    r = requests.post('http://www.freightcom.com/rpc2', data=xml, headers=headers, verify=False)

    with open('quotes.xml', 'wb') as f:
       f.write(r.content)
    f.close()

    tree = ET.parse('quotes.xml')
    root = tree.getroot()
    for quote in root.iter('{http://www.freightcom.net/xml/XMLSchema}Quote'):
        attr = quote.attrib
        print attr["baseCharge"]
